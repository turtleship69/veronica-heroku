import discord
from discord.ext import commands
import json
import os
import psutil
import random
import requests
import datetime

from openpyxl import load_workbook

wb = load_workbook("dadsaysjokes_user_tweets.xlsx")  # Work Book
ws = wb['tweets']  # Work Sheet
column = ws['C']  # Column
column_list = [column[x].value for x in range(len(column))]

import keep_alive

keep_alive.keep_alive()

activity = discord.Game(name="Take On Me")

client = commands.Bot(command_prefix="\\", activity=activity)


@client.event
async def on_ready():
    print("Bot is ready")


@client.command()
async def dad(ctx):
    joke = column_list[random.randint(0, (len(column_list) - 1))]
    await ctx.send(joke)


@client.command()
async def hello(ctx):
    await ctx.send("Hi")


@client.command()
async def echo(ctx, *, message):
    await ctx.author.send(message)


# def getMeme():
#     URL = "https://meme-api.herokuapp.com/gimme"
#     response = requests.get(URL)
#     if response.status_code == 200:
#         data = response.json()
#         if data['nsfw'] == False:
#             return data['url']
#         else:
#             return 'try again'
#     else:
#         return "HTTP error "+str(response.status_code)

# @client.command()
# async def meme(ctx):
#     await ctx.send(getMeme())


@client.command()
async def meme(ctx):
    content = requests.get("https://meme-api.herokuapp.com/gimme").text
    data = json.loads(content, )
    mTitle = data['title']
    mUrl = data['url']
    mPostLink = data['postLink']
    meme = discord.Embed(title=f"{mTitle}",
                         Color=discord.Color.random(),
                         url=mPostLink).set_image(url=f"{mUrl}")
    meme.set_footer(text=f"{mPostLink}")
    await ctx.reply(embed=meme)


@client.command(aliases=['kitty', 'cats', 'cute', 'pussy'])
async def cat(ctx):
    content = requests.get("https://meme-api.herokuapp.com/gimme/cat").text
    data = json.loads(content, )
    mTitle = data['title']
    mUrl = data['url']
    mPostLink = data['postLink']
    meme = discord.Embed(title=f"{mTitle}",
                         Color=discord.Color.random(),
                         url=mPostLink).set_image(url=f"{mUrl}")
    meme.set_footer(text=f"{mPostLink}")
    await ctx.reply(embed=meme)


#####################################################
#economy start
mainshop = [{
    "name": "Watch",
    "price": 100,
    "description": "Time"
}, {
    "name": "Laptop",
    "price": 1000,
    "description": "Work"
}, {
    "name": "PC",
    "price": 10000,
    "description": "Gaming"
}, {
    "name": "Ferrari",
    "price": 99999,
    "description": "Sports Car"
}]


@client.command()
async def ping(ctx):
    await ctx.reply(f'Pong')  # {ctx.author.mention}


def __init__(self, client):
    self.client = client


@client.command(aliases=['python', 'botinfo'])
async def bot(ctx):
    values = psutil.virtual_memory()
    val2 = values.available * 0.001
    val3 = val2 * 0.001
    val4 = val3 * 0.001
    values2 = psutil.virtual_memory()
    value21 = values2.total
    values22 = value21 * 0.001
    values23 = values22 * 0.001
    values24 = values23 * 0.001
    embedve = discord.Embed(title="Bot Info", description=None, color=0x9370DB)
    #    embedve.add_field(
    #        name="Bot Latency", value=f"Bot latency - {round(self.client.latency * 1000)} ms", inline=False)
    embedve.add_field(
        name='Hosting Stats',
        value=f'Cpu usage- {psutil.cpu_percent(1)}%'
        f'\n(Actual Cpu Usage May Differ)'
        f'\n'
        f'\nNumber OF Cores - {psutil.cpu_count()} '
        f'\nNumber of Physical Cores- {psutil.cpu_count(logical=False)}'
        f'\n'
        f'\nTotal ram- {round(values24, 2)} GB'
        f'\nAvailable Ram - {round(val4, 2)} GB')
    await ctx.send(embed=embedve)


client.remove_command("help")


@client.command()
async def help(ctx):
    embedvar = discord.Embed(title="Help Commands",
                             description=None,
                             color=0x00ff00)
    embedvar.add_field(name='\\bot', value='To see bot info', inline=False)
    embedvar.add_field(name='\\balance / \\bal',
                       value='To see your balance',
                       inline=False)
    embedvar.add_field(name='\\beg', value='To beg some money', inline=False)
    embedvar.add_field(name='\\deposit',
                       value='To deposit money in bank',
                       inline=False)
    embedvar.add_field(name='\\withdraw',
                       value='To withdraw money from bank',
                       inline=False)
    embedvar.add_field(name='\\send',
                       value='Send money to someone',
                       inline=False)
    embedvar.add_field(name='\\rob',
                       value='Rob some random money ',
                       inline=False)
    embedvar.add_field(name='\\slots', value='To bet some money', inline=False)
    embedvar.add_field(name='\\shop', value='To view shop', inline=False)
    embedvar.add_field(name='\\buy', value='To, buy an item', inline=False)
    embedvar.add_field(name='\\sell', value='To sell an item', inline=False)
    embedvar.add_field(name='\\bag',
                       value='To view your shopping cart',
                       inline=False)
    embedvar.add_field(name='\\lb', value='To view leaderboard', inline=False)
    await ctx.send(embed=embedvar)


@client.command(aliases=['bal'])
async def balance(ctx):
    await open_account(ctx.author)
    user = ctx.author

    users = await get_bank_data()

    wallet_amt = users[str(user.id)]["wallet"]
    bank_amt = users[str(user.id)]["bank"]

    em = discord.Embed(title=f'{ctx.author.name} Balance',
                       color=discord.Color.red())
    em.add_field(name="Wallet Balance", value=wallet_amt)
    em.add_field(name='Bank Balance', value=bank_amt)
    await ctx.send(embed=em)


@client.command()
async def beg(ctx):
    await open_account(ctx.author)
    user = ctx.author

    users = await get_bank_data()

    earnings = random.randrange(101)

    await ctx.send(f'{ctx.author.mention} Got {earnings} coins!!')

    users[str(user.id)]["wallet"] += earnings

    with open("mainbank.json", 'w') as f:
        json.dump(users, f)


@client.command(aliases=['wd'])
async def withdraw(ctx, amount=None):
    await open_account(ctx.author)
    if amount == None:
        await ctx.send("Please enter the amount")
        return

    bal = await update_bank(ctx.author)

    amount = int(amount)

    if amount > bal[1]:
        await ctx.send('You do not have sufficient balance')
        return
    if amount < 0:
        await ctx.send('Amount must be positive!')
        return

    await update_bank(ctx.author, amount)
    await update_bank(ctx.author, -1 * amount, 'bank')
    await ctx.send(f'{ctx.author.mention} You withdrew {amount} coins')


@client.command(aliases=['dp'])
async def deposit(ctx, amount=None):
    await open_account(ctx.author)
    if amount == None:
        await ctx.send("Please enter the amount")
        return

    bal = await update_bank(ctx.author)

    amount = int(amount)

    if amount > bal[0]:
        await ctx.send('You do not have sufficient balance')
        return
    if amount < 0:
        await ctx.send('Amount must be positive!')
        return

    await update_bank(ctx.author, -1 * amount)
    await update_bank(ctx.author, amount, 'bank')
    await ctx.send(f'{ctx.author.mention} You deposited {amount} coins')


@client.command(aliases=['sm'])
async def send(ctx, member: discord.Member, amount=None):
    await open_account(ctx.author)
    await open_account(member)
    if amount == None:
        await ctx.send("Please enter the amount")
        return

    bal = await update_bank(ctx.author)
    if amount == 'all':
        amount = bal[0]

    amount = int(amount)

    if amount > bal[0]:
        await ctx.send('You do not have sufficient balance')
        return
    if amount < 0:
        await ctx.send('Amount must be positive!')
        return

    await update_bank(ctx.author, -1 * amount, 'bank')
    await update_bank(member, amount, 'bank')
    await ctx.send(f'{ctx.author.mention} You gave {member} {amount} coins')


@client.command(aliases=['rb'])
async def rob(ctx, member: discord.Member):
    await open_account(ctx.author)
    await open_account(member)
    bal = await update_bank(member)

    if bal[0] < 100:
        await ctx.send('It is useless to rob him :(')
        return

    earning = random.randrange(0, bal[0])

    await update_bank(ctx.author, earning)
    await update_bank(member, -1 * earning)
    await ctx.send(
        f'{ctx.author.mention} You robbed {member} and got {earning} coins')


@client.command()
async def slots(ctx, amount=None):
    await open_account(ctx.author)
    if amount == None:
        await ctx.send("Please enter the amount")
        return

    bal = await update_bank(ctx.author)

    amount = int(amount)

    if amount > bal[0]:
        await ctx.send('You do not have sufficient balance')
        return
    if amount < 0:
        await ctx.send('Amount must be positive!')
        return
    final = []
    for i in range(3):
        a = random.choice(['X', 'O', 'Q'])

        final.append(a)

    await ctx.send(str(final))

    if final[0] == final[1] or final[1] == final[2] or final[0] == final[2]:
        await update_bank(ctx.author, 2 * amount)
        await ctx.send(f'You won :) {ctx.author.mention}')
    else:
        await update_bank(ctx.author, -1 * amount)
        await ctx.send(f'You lose :( {ctx.author.mention}')


@client.command()
async def shop(ctx):
    em = discord.Embed(title="Shop")

    for item in mainshop:
        name = item["name"]
        price = item["price"]
        desc = item["description"]
        em.add_field(name=name, value=f"${price} | {desc}")

    await ctx.send(embed=em)


@client.command()
async def buy(ctx, item, amount=1):
    await open_account(ctx.author)

    res = await buy_this(ctx.author, item, amount)

    if not res[0]:
        if res[1] == 1:
            await ctx.send("That Object isn't there!")
            return
        if res[1] == 2:
            await ctx.send(
                f"You don't have enough money in your wallet to buy {amount} {item}"
            )
            return

    await ctx.send(f"You just bought {amount} {item}")


@client.command()
async def bag(ctx):
    await open_account(ctx.author)
    user = ctx.author
    users = await get_bank_data()

    try:
        bag = users[str(user.id)]["bag"]
    except:
        bag = []

    em = discord.Embed(title="Bag")
    for item in bag:
        name = item["item"]
        amount = item["amount"]

        em.add_field(name=name, value=amount)

    await ctx.send(embed=em)


async def buy_this(user, item_name, amount):
    item_name = item_name.lower()
    name_ = None
    for item in mainshop:
        name = item["name"].lower()
        if name == item_name:
            name_ = name
            price = item["price"]
            break

    if name_ == None:
        return [False, 1]

    cost = price * amount

    users = await get_bank_data()

    bal = await update_bank(user)

    if bal[0] < cost:
        return [False, 2]

    try:
        index = 0
        t = None
        for thing in users[str(user.id)]["bag"]:
            n = thing["item"]
            if n == item_name:
                old_amt = thing["amount"]
                new_amt = old_amt + amount
                users[str(user.id)]["bag"][index]["amount"] = new_amt
                t = 1
                break
            index += 1
        if t == None:
            obj = {"item": item_name, "amount": amount}
            users[str(user.id)]["bag"].append(obj)
    except:
        obj = {"item": item_name, "amount": amount}
        users[str(user.id)]["bag"] = [obj]

    with open("mainbank.json", "w") as f:
        json.dump(users, f)

    await update_bank(user, cost * -1, "wallet")

    return [True, "Worked"]


@client.command()
async def sell(ctx, item, amount=1):
    await open_account(ctx.author)

    res = await sell_this(ctx.author, item, amount)

    if not res[0]:
        if res[1] == 1:
            await ctx.send("That Object isn't there!")
            return
        if res[1] == 2:
            await ctx.send(f"You don't have {amount} {item} in your bag.")
            return
        if res[1] == 3:
            await ctx.send(f"You don't have {item} in your bag.")
            return

    await ctx.send(f"You just sold {amount} {item}.")


async def sell_this(user, item_name, amount, price=None):
    item_name = item_name.lower()
    name_ = None
    for item in mainshop:
        name = item["name"].lower()
        if name == item_name:
            name_ = name
            if price == None:
                price = 0.7 * item["price"]
            break

    if name_ == None:
        return [False, 1]

    cost = price * amount

    users = await get_bank_data()

    bal = await update_bank(user)

    try:
        index = 0
        t = None
        for thing in users[str(user.id)]["bag"]:
            n = thing["item"]
            if n == item_name:
                old_amt = thing["amount"]
                new_amt = old_amt - amount
                if new_amt < 0:
                    return [False, 2]
                users[str(user.id)]["bag"][index]["amount"] = new_amt
                t = 1
                break
            index += 1
        if t == None:
            return [False, 3]
    except:
        return [False, 3]

    with open("mainbank.json", "w") as f:
        json.dump(users, f)

    await update_bank(user, cost, "wallet")

    return [True, "Worked"]


@client.command(aliases=["lb"])
async def leaderboard(ctx, x=1):
    users = await get_bank_data()
    leader_board = {}
    total = []
    for user in users:
        name = int(user)
        total_amount = users[user]["wallet"] + users[user]["bank"]
        leader_board[total_amount] = name
        total.append(total_amount)

    total = sorted(total, reverse=True)

    em = discord.Embed(
        title=f"Top {x} Richest People",
        description=
        "This is decided on the basis of raw money in the bank and wallet",
        color=discord.Color(0xfa43ee))
    index = 1
    for amt in total:
        id_ = leader_board[amt]
        member = client.get_user(id_)
        name = member.name
        em.add_field(name=f"{index}. {name}", value=f"{amt}", inline=False)
        if index == x:
            break
        else:
            index += 1

    await ctx.send(embed=em)


async def open_account(user):

    users = await get_bank_data()

    if str(user.id) in users:
        return False
    else:
        users[str(user.id)] = {}
        users[str(user.id)]["wallet"] = 0
        users[str(user.id)]["bank"] = 0

    with open('mainbank.json', 'w') as f:
        json.dump(users, f)

    return True


async def get_bank_data():
    with open('mainbank.json', 'r') as f:
        users = json.load(f)

    return users


async def update_bank(user, change=0, mode='wallet'):
    users = await get_bank_data()

    users[str(user.id)][mode] += change

    with open('mainbank.json', 'w') as f:
        json.dump(users, f)
    bal = users[str(user.id)]['wallet'], users[str(user.id)]['bank']
    return bal


#economy end
#####################################################
def check(message: discord.Message) -> bool:
    if message.author.id == ctx.author.id:
        return True
    return False


@client.command()
async def murder(ctx):
    await ctx.reply("Are you sure?")
    msg = await client.wait_for('message', timeout=15.0)
    if msg.content.lower() == 'yes':
        await msg.reply("Cya")
        import time, sys
        time.sleep(2)
        #sys.exit()
        os.system("killall -KILL prybar-python3")
        #time.sleep(0.5)
        #os.system("pkill python")

    message = await client.wait_for('message', check=check)
    await send("f")
    #if ctx.author.id == 520644933074550784:


API = os.getenv("API")
client.run(API)

#You touch my bot I end your life by several different metrics
